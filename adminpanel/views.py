# adminpanel/views.py
from rest_framework import generics, permissions, status
from rest_framework.response import Response
from rest_framework.views import APIView
from django.shortcuts import get_object_or_404
from django.db.models import Q
from django.utils import timezone
from django.http import HttpResponse
import csv
import json
from datetime import timedelta

from .models import AdminLog, SystemSetting, AdminNotification
from .serializers import (
    AdminLogSerializer, SystemSettingSerializer,
    AdminNotificationSerializer, MarkNotificationReadSerializer,
    ExportDataSerializer, ImportDataSerializer
)
from accounts.models import User
from orders.models import Order
from products.models import Product
from analytics.models import SalesReport


class AdminDashboardView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request, *args, **kwargs):
        # Time ranges
        today = timezone.now().date()
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)

        # Order metrics
        total_orders = Order.objects.count()
        recent_orders = Order.objects.filter(
            created_at__date__gte=week_ago
        ).count()
        pending_orders = Order.objects.filter(
            status='pending'
        ).count()

        # Revenue metrics
        total_revenue = Order.objects.filter(
            payment_status='paid'
        ).aggregate(
            total=Sum('total')
        )['total'] or 0
        recent_revenue = Order.objects.filter(
            payment_status='paid',
            created_at__date__gte=month_ago
        ).aggregate(
            total=Sum('total')
        )['total'] or 0

        # User metrics
        total_users = User.objects.count()
        new_users = User.objects.filter(
            date_joined__date__gte=month_ago
        ).count()

        # Product metrics
        total_products = Product.objects.count()
        low_stock_products = Product.objects.filter(
            stock__lt=10
        ).count()

        # Recent activities
        recent_activities = AdminLog.objects.filter(
            created_at__gte=timezone.now() - timedelta(days=7)
        ).order_by('-created_at')[:10]

        # Recent notifications
        recent_notifications = AdminNotification.objects.filter(
            user=request.user,
            is_read=False
        ).order_by('-created_at')[:5]

        data = {
            'metrics': {
                'orders': {
                    'total': total_orders,
                    'recent': recent_orders,
                    'pending': pending_orders,
                },
                'revenue': {
                    'total': float(total_revenue),
                    'recent': float(recent_revenue),
                },
                'users': {
                    'total': total_users,
                    'new': new_users,
                },
                'products': {
                    'total': total_products,
                    'low_stock': low_stock_products,
                }
            },
            'recent_activities': AdminLogSerializer(
                recent_activities,
                many=True
            ).data,
            'recent_notifications': AdminNotificationSerializer(
                recent_notifications,
                many=True
            ).data,
        }

        return Response(data)


class SystemSettingListView(generics.ListCreateAPIView):
    serializer_class = SystemSettingSerializer
    permission_classes = [permissions.IsAdminUser]
    queryset = SystemSetting.objects.all()
    search_fields = ['key', 'description']
    filterset_fields = ['setting_type', 'is_public']


class SystemSettingDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = SystemSettingSerializer
    permission_classes = [permissions.IsAdminUser]
    queryset = SystemSetting.objects.all()
    lookup_field = 'key'


class AdminLogListView(generics.ListAPIView):
    serializer_class = AdminLogSerializer
    permission_classes = [permissions.IsAdminUser]
    queryset = AdminLog.objects.all()
    filter_backends = []
    pagination_class = None

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filter by action type if provided
        action_type = self.request.query_params.get('action_type')
        if action_type:
            queryset = queryset.filter(action_type=action_type)

        # Filter by model name if provided
        model_name = self.request.query_params.get('model_name')
        if model_name:
            queryset = queryset.filter(model_name__iexact=model_name)

        # Filter by date range if provided
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date and end_date:
            queryset = queryset.filter(
                created_at__date__gte=start_date,
                created_at__date__lte=end_date
            )

        return queryset.order_by('-created_at')[:100]  # Limit to 100 most recent


class AdminNotificationListView(generics.ListAPIView):
    serializer_class = AdminNotificationSerializer
    permission_classes = [permissions.IsAdminUser]

    def get_queryset(self):
        return AdminNotification.objects.filter(
            user=self.request.user
        ).order_by('-created_at')


class MarkNotificationReadView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def post(self, request, *args, **kwargs):
        serializer = MarkNotificationReadSerializer(
            data=request.data,
            context={'request': request}
        )
        serializer.is_valid(raise_exception=True)

        AdminNotification.objects.filter(
            id__in=serializer.validated_data['ids'],
            user=request.user
        ).update(is_read=True, read_at=timezone.now())

        return Response(
            {'status': 'success', 'message': 'Notifications marked as read'},
            status=status.HTTP_200_OK
        )


class MarkAllNotificationsReadView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def post(self, request, *args, **kwargs):
        updated = AdminNotification.objects.filter(
            user=request.user,
            is_read=False
        ).update(is_read=True, read_at=timezone.now())

        return Response(
            {
                'status': 'success',
                'message': f'{updated} notifications marked as read'
            },
            status=status.HTTP_200_OK
        )


class ExportDataView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def post(self, request, *args, **kwargs):
        serializer = ExportDataSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        model_name = serializer.validated_data['model_name']
        export_format = serializer.validated_data['format']
        filters = serializer.validated_data.get('filters', {})

        # Get model class
        model = self._get_model_class(model_name)
        if not model:
            return Response(
                {'error': 'Invalid model name'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Apply filters
        queryset = model.objects.all()
        if filters:
            queryset = queryset.filter(**filters)

        # Export data
        if export_format == 'json':
            return self._export_json(queryset, model_name)
        elif export_format == 'csv':
            return self._export_csv(queryset, model_name)
        elif export_format == 'xlsx':
            return self._export_excel(queryset, model_name)

        return Response(
            {'error': 'Unsupported export format'},
            status=status.HTTP_400_BAD_REQUEST
        )

    def _get_model_class(self, model_name):
        """Get model class from string name"""
        from django.apps import apps
        try:
            return apps.get_model(model_name)
        except LookupError:
            return None

    def _export_json(self, queryset, model_name):
        """Export data as JSON"""
        from django.core import serializers
        data = serializers.serialize('json', queryset)
        response = HttpResponse(data, content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="{model_name}_export.json"'
        return response

    def _export_csv(self, queryset, model_name):
        """Export data as CSV"""
        import csv
        from io import StringIO

        # Get field names
        field_names = [field.name for field in queryset.model._meta.fields]

        # Create CSV writer
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(field_names)

        # Write data rows
        for obj in queryset:
            row = [getattr(obj, field) for field in field_names]
            writer.writerow(row)

        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{model_name}_export.csv"'
        return response

    def _export_excel(self, queryset, model_name):
        """Export data as Excel"""
        import openpyxl
        from openpyxl.utils import get_column_letter

        # Create workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = model_name

        # Get field names
        field_names = [field.name for field in queryset.model._meta.fields]

        # Write headers
        for col_num, field_name in enumerate(field_names, 1):
            col_letter = get_column_letter(col_num)
            worksheet[f'{col_letter}1'] = field_name

        # Write data rows
        for row_num, obj in enumerate(queryset, 2):
            for col_num, field_name in enumerate(field_names, 1):
                col_letter = get_column_letter(col_num)
                worksheet[f'{col_letter}{row_num}'] = getattr(obj, field_name)

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{model_name}_export.xlsx"'
        workbook.save(response)
        return response


class ImportDataView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def post(self, request, *args, **kwargs):
        serializer = ImportDataSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        model_name = serializer.validated_data['model_name']
        import_file = serializer.validated_data['file']
        import_format = serializer.validated_data['format']
        update_existing = serializer.validated_data['update_existing']

        # Get model class
        model = self._get_model_class(model_name)
        if not model:
            return Response(
                {'error': 'Invalid model name'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            if import_format == 'json':
                result = self._import_json(model, import_file, update_existing)
            elif import_format == 'csv':
                result = self._import_csv(model, import_file, update_existing)
            else:
                return Response(
                    {'error': 'Unsupported import format'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            return Response({
                'status': 'success',
                'created': result['created'],
                'updated': result['updated'],
                'skipped': result['skipped'],
                'errors': result['errors']
            })
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    def _get_model_class(self, model_name):
        """Get model class from string name"""
        from django.apps import apps
        try:
            return apps.get_model(model_name)
        except LookupError:
            return None

    def _import_json(self, model, import_file, update_existing):
        """Import data from JSON"""
        import json
        from django.core import serializers

        data = json.load(import_file)
        result = {
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': []
        }

        for item in data:
            try:
                if update_existing and 'pk' in item:
                    instance, created = model.objects.update_or_create(
                        pk=item['pk'],
                        defaults=item['fields']
                    )
                    if created:
                        result['created'] += 1
                    else:
                        result['updated'] += 1
                else:
                    model.objects.create(**item['fields'])
                    result['created'] += 1
            except Exception as e:
                result['errors'].append(str(e))
                result['skipped'] += 1

        return result

    def _import_csv(self, model, import_file, update_existing):
        """Import data from CSV"""
        import csv
        from io import TextIOWrapper

        reader = csv.DictReader(TextIOWrapper(import_file))
        field_names = [field.name for field in model._meta.fields]

        result = {
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': []
        }

        for row in reader:
            try:
                # Filter row data to only include model fields
                filtered_data = {
                    k: v for k, v in row.items()
                    if k in field_names
                }

                if update_existing and 'id' in filtered_data:
                    instance, created = model.objects.update_or_create(
                        id=filtered_data['id'],
                        defaults=filtered_data
                    )
                    if created:
                        result['created'] += 1
                    else:
                        result['updated'] += 1
                else:
                    model.objects.create(**filtered_data)
                    result['created'] += 1
            except Exception as e:
                result['errors'].append(str(e))
                result['skipped'] += 1

        return result


class GenerateSalesReportView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def post(self, request, *args, **kwargs):
        from analytics.models import SalesReport
        from datetime import date

        # Default to last month if no dates provided
        today = date.today()
        first_day = today.replace(day=1)
        last_month = first_day - timedelta(days=1)
        start_date = last_month.replace(day=1)
        end_date = last_month

        # Create or update report
        report, created = SalesReport.objects.get_or_create(
            period_type='monthly',
            start_date=start_date,
            end_date=end_date,
            defaults={
                'period_type': 'monthly',
                'start_date': start_date,
                'end_date': end_date,
            }
        )

        # Calculate metrics
        report.calculate_metrics()
        report.save()

        return Response(
            {
                'status': 'success',
                'message': f"{report.get_period_type_display()} report generated",
                'report': {
                    'start_date': report.start_date,
                    'end_date': report.end_date,
                    'total_orders': report.total_orders,
                    'total_revenue': report.total_revenue
                }
            },
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK
        )